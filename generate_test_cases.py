"""
Generate comprehensive Dev Test Cases for B2B Cab Portal.

Output: C:\\Users\\Pranav\\Downloads\\B2B_Cab_Dev_Test_Cases.xlsx

Structure:
  - Summary sheet (totals, sign-off)
  - Test Data sheet (credentials, sample data)
  - 18 module sheets with ~170 test cases total
  - Standard columns: ID, Module, Title, Type, Severity, Pre-conditions,
                      Test Data, Steps, Expected Result, Status, Tester, Date, Notes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


# ─────────── COLORS ───────────
PRIMARY = "0EA5E9"           # sky-500
DARK_BG = "0F172A"           # slate-900
DARK_TEXT = "FFFFFF"
HEADER_BG = "0EA5E9"         # sky-500 for column headers
HEADER_TEXT = "FFFFFF"
ALT_ROW = "F8FAFC"           # slate-50
BORDER_COLOR = "CBD5E1"      # slate-300
SECTION_BG = "E0F2FE"        # sky-100

# Severity colors
SEV_CRITICAL = "FECACA"      # red-200
SEV_HIGH = "FED7AA"          # orange-200
SEV_MEDIUM = "FEF08A"        # yellow-200
SEV_LOW = "E5E7EB"           # gray-200

# Status colors (for conditional formatting)
STATUS_PASS = "BBF7D0"       # green-200
STATUS_FAIL = "FECACA"       # red-200
STATUS_BLOCK = "DDD6FE"      # violet-200
STATUS_PENDING = "F1F5F9"    # slate-100


# ─────────── BORDER ───────────
thin = Side(border_style="thin", color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─────────── COLUMN HEADERS ───────────
HEADERS = [
    ("Test ID", 12),
    ("Module", 16),
    ("Title", 45),
    ("Type", 12),
    ("Severity", 12),
    ("Pre-conditions", 35),
    ("Test Data", 30),
    ("Steps", 55),
    ("Expected Result", 50),
    ("Status", 12),
    ("Tested By", 14),
    ("Date", 12),
    ("Notes", 25),
]


def style_header_row(ws, row=1):
    """Apply header styling to a row."""
    for col_idx, (h, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(bold=True, color=HEADER_TEXT, size=10, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 30


def add_test_case(ws, row, tc):
    """Write one test case row.
    tc = (id, module, title, type, severity, preconditions, test_data, steps, expected, notes)
    """
    sev_fill = {
        "Critical": SEV_CRITICAL,
        "High": SEV_HIGH,
        "Medium": SEV_MEDIUM,
        "Low": SEV_LOW,
    }.get(tc[4], SEV_LOW)

    values = [
        tc[0],          # ID
        tc[1],          # Module
        tc[2],          # Title
        tc[3],          # Type
        tc[4],          # Severity
        tc[5],          # Pre-conditions
        tc[6],          # Test Data
        tc[7],          # Steps
        tc[8],          # Expected Result
        "Pending",      # Status
        "",             # Tested By
        "",             # Date
        tc[9] if len(tc) > 9 else "",  # Notes
    ]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(size=9, name='Calibri', color="0F172A")
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border
        # Severity column gets fill
        if col_idx == 5:
            cell.fill = PatternFill("solid", fgColor=sev_fill)
            cell.font = Font(size=9, name='Calibri', bold=True, color="0F172A")
        # ID column bold
        if col_idx == 1:
            cell.font = Font(size=9, name='Consolas', bold=True, color="0F172A")
        # Status column - center align
        if col_idx == 10:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill("solid", fgColor=STATUS_PENDING)


def add_module_sheet(wb, name, test_cases):
    """Create a sheet for a module with header + test cases."""
    ws = wb.create_sheet(title=name)
    style_header_row(ws, row=1)
    ws.freeze_panes = "A2"

    for i, tc in enumerate(test_cases, start=2):
        add_test_case(ws, i, tc)
        # Set row height to fit content
        ws.row_dimensions[i].height = 60

    # Add data validation for Status column
    dv = DataValidation(type="list", formula1='"Pending,Pass,Fail,Blocked,Skip"', allow_blank=True)
    dv.add(f"J2:J{len(test_cases) + 1}")
    ws.add_data_validation(dv)

    return ws


# ════════════════════════════════════════════════════════════════
# TEST CASES BY MODULE
# ════════════════════════════════════════════════════════════════


# ─── 1. AUTHENTICATION ───────────────────────────────────────────
AUTH_TESTS = [
    ("TC-AUTH-001", "Auth", "Successful login with valid email + password", "Functional", "Critical",
     "Internet active, valid agent account exists",
     "Email: bincy.joseph@savaari.com\nPwd: aMuWysE@YgAVa5aPagYR",
     "1. Navigate to /login\n2. Enter email\n3. Enter password\n4. Click Login button",
     "Spinner shows briefly, redirects to /dashboard, JWT stored in localStorage as loginUserToken, user name shows in header",
     ""),

    ("TC-AUTH-002", "Auth", "Login fails with wrong password", "Negative", "High",
     "Valid email, wrong password",
     "Email: bincy.joseph@savaari.com\nPwd: wrongpass123",
     "1. Open /login\n2. Enter valid email\n3. Enter wrong password\n4. Click Login",
     "Error toast/message: 'Invalid credentials' or similar. No redirect. localStorage empty",
     ""),

    ("TC-AUTH-003", "Auth", "Login fails with non-existent email", "Negative", "High",
     "Email not registered",
     "Email: doesnotexist@test.com\nPwd: anything123",
     "1. Open /login\n2. Enter random email\n3. Enter password\n4. Click Login",
     "Error message shown. No redirect",
     ""),

    ("TC-AUTH-004", "Auth", "Login button disabled when fields empty", "Validation", "Medium",
     "Fresh login page",
     "—",
     "1. Open /login\n2. Leave fields blank\n3. Observe Login button",
     "Login button is disabled or shows validation errors when clicked",
     ""),

    ("TC-AUTH-005", "Auth", "Email format validation", "Validation", "Medium",
     "Login page open",
     "Email: not-an-email\nPwd: anything",
     "1. Type 'not-an-email' in email field\n2. Tab out\n3. Observe field state",
     "Validation error: 'Invalid email format' shown below field",
     ""),

    ("TC-AUTH-006", "Auth", "Password visibility toggle", "UI", "Low",
     "Login page open",
     "Pwd: testpass123",
     "1. Type password in field\n2. Click eye icon\n3. Click eye icon again",
     "First click shows plaintext, second click hides as dots",
     ""),

    ("TC-AUTH-007", "Auth", "Logout clears session", "Functional", "Critical",
     "User logged in",
     "—",
     "1. Login successfully\n2. Click profile menu\n3. Click Logout",
     "Redirected to /login. localStorage cleared (no SavaariToken/loginUserToken). Cannot access /dashboard via direct URL",
     ""),

    ("TC-AUTH-008", "Auth", "Direct URL access to protected route without auth", "Security", "Critical",
     "Not logged in",
     "URL: /dashboard",
     "1. Logout if logged in\n2. Type /dashboard in URL bar\n3. Press Enter",
     "Redirected to /login (authGuard blocks). No flash of dashboard content",
     ""),

    ("TC-AUTH-009", "Auth", "Already logged-in user visiting /login", "Functional", "Medium",
     "User logged in",
     "URL: /login",
     "1. Login successfully\n2. Type /login in URL bar\n3. Press Enter",
     "Redirected to /dashboard automatically (guestGuard blocks). Login form not shown",
     ""),

    ("TC-AUTH-010", "Auth", "Token persists across page refresh", "Functional", "High",
     "Logged in user",
     "—",
     "1. Login successfully\n2. Press F5 to refresh\n3. Observe app state",
     "User stays logged in. Dashboard reloads with same user. No re-auth needed",
     ""),

    ("TC-AUTH-011", "Auth", "SQL injection attempt in email field", "Security", "High",
     "Login page open",
     "Email: ' OR '1'='1\nPwd: anything",
     "1. Type SQL injection string in email\n2. Click Login",
     "Login fails with normal error. No SQL error leaked. No unauthorized access",
     ""),

    ("TC-AUTH-012", "Auth", "Multiple failed logins do not lock UI", "Functional", "Medium",
     "Login page open",
     "Email: any\nPwd: wrong (×5)",
     "1. Try login with wrong password 5 times in a row",
     "Each attempt shows error. Button remains clickable. No frontend crash. Backend may rate-limit",
     ""),
]


# ─── 2. DASHBOARD ────────────────────────────────────────────────
DASHBOARD_TESTS = [
    ("TC-DASH-001", "Dashboard", "Dashboard loads after login with all elements", "Functional", "Critical",
     "Logged in user",
     "—",
     "1. Login\n2. Land on /dashboard\n3. Observe page",
     "Header with user name + wallet balance, 4 trip type tabs, search form card, recent bookings widget visible",
     ""),

    ("TC-DASH-002", "Dashboard", "Switch between 4 trip type tabs", "Functional", "Critical",
     "On /dashboard",
     "—",
     "1. Click 'One Way'\n2. Click 'Round Trip'\n3. Click 'Local'\n4. Click 'Airport'",
     "Each tab shows correct fields. Active tab highlighted. Form fields update accordingly",
     ""),

    ("TC-DASH-003", "Dashboard", "One Way tab shows correct fields", "UI", "High",
     "Click One Way tab",
     "—",
     "1. Click One Way tab\n2. Observe fields",
     "Fields visible: From City, To City, Pickup Date, Pickup Time, Search button. No duration field",
     ""),

    ("TC-DASH-004", "Dashboard", "Round Trip tab shows return date", "UI", "High",
     "Click Round Trip tab",
     "—",
     "1. Click Round Trip\n2. Observe fields",
     "Fields visible: From, To, Pickup Date, Pickup Time, Return Date, Return Time, Add Stop button",
     ""),

    ("TC-DASH-005", "Dashboard", "Local tab shows package selector", "UI", "High",
     "Click Local tab",
     "—",
     "1. Click Local\n2. Observe fields",
     "Fields visible: City, Pickup Date, Pickup Time. No drop city field. Duration handled later on select-car",
     ""),

    ("TC-DASH-006", "Dashboard", "Airport tab default sub-type is 'drop'", "UI", "High",
     "Click Airport tab",
     "—",
     "1. Click Airport tab\n2. Observe sub trip type field",
     "Trip Type dropdown defaults to 'Drop to Airport'. Field width 240px. No 'Select type' blank",
     "Recently fixed: was blank earlier"),

    ("TC-DASH-007", "Dashboard", "City autocomplete triggers after 3 chars", "Functional", "High",
     "Form open",
     "Type: 'Mum'",
     "1. Click From City\n2. Type 'Mum' slowly\n3. Observe dropdown",
     "Autocomplete dropdown appears with cities matching (Mumbai, Mumbra, etc.). Loading spinner during fetch",
     ""),

    ("TC-DASH-008", "Dashboard", "City swap button swaps from/to", "Functional", "Medium",
     "Both cities selected",
     "From: Mumbai\nTo: Pune",
     "1. Select From: Mumbai, To: Pune\n2. Click swap button (↔)",
     "Cities swap. From: Pune, To: Mumbai. City IDs also correctly swapped (verify in network call)",
     ""),

    ("TC-DASH-009", "Dashboard", "Past date cannot be selected", "Validation", "High",
     "Form open",
     "Date: yesterday",
     "1. Click date picker\n2. Try to select yesterday\n3. Observe",
     "Past dates greyed out / not clickable. Min date = today",
     ""),

    ("TC-DASH-010", "Dashboard", "Time picker shows 15-min intervals", "UI", "Medium",
     "Form open",
     "—",
     "1. Click time picker\n2. Open hours/minutes dropdown",
     "Minutes dropdown shows 00, 15, 30, 45 only (15-min intervals)",
     ""),

    ("TC-DASH-011", "Dashboard", "Recent bookings widget loads", "Functional", "Medium",
     "Logged in with past bookings",
     "—",
     "1. Land on /dashboard\n2. Scroll to recent bookings",
     "Widget shows last 3-5 bookings with status badges. If no bookings, shows empty state",
     ""),

    ("TC-DASH-012", "Dashboard", "Wallet balance refreshes on dashboard", "Functional", "High",
     "Logged in user",
     "—",
     "1. Land on /dashboard\n2. Observe wallet balance in header",
     "Wallet balance shown in INR. Matches /wallet page balance. Updates after top-up or booking",
     ""),
]


# ─── 3. OUTSTATION ONE WAY ──────────────────────────────────────
ONEWAY_TESTS = [
    ("TC-OW-001", "One Way", "Complete one-way booking from search to confirmation", "Functional", "Critical",
     "Logged in, wallet has balance",
     "From: Mumbai\nTo: Pune\nDate: tomorrow\nTime: 10:00",
     "1. Select One Way tab\n2. Fill from/to/date/time\n3. Click Search\n4. Select a car\n5. Fill passenger info\n6. Pay via wallet Option 1 (25%)\n7. Confirm",
     "Booking created. Confirmation page shows booking ID, fare breakdown, timeline. Wallet deducted",
     "End-to-end critical path"),

    ("TC-OW-002", "One Way", "Source = destination should show error", "Negative", "High",
     "Form open",
     "From: Mumbai\nTo: Mumbai",
     "1. Select Mumbai for both From and To\n2. Click Search",
     "Validation error: 'Pickup and drop cannot be same' or button disabled",
     ""),

    ("TC-OW-003", "One Way", "Search returns car list", "Functional", "Critical",
     "Valid form filled",
     "Mumbai → Pune tomorrow",
     "1. Fill valid form\n2. Click Search\n3. Wait for /select-car page",
     "Car list loads with at least 3 cars. Each car shows price, KMs, name, image",
     ""),

    ("TC-OW-004", "One Way", "Drop address field is shown for One Way", "UI", "High",
     "On booking details page (One Way)",
     "Booking via One Way",
     "1. Reach /booking page after car select\n2. Look at Drop Address field",
     "Drop Address field visible and required for One Way trips",
     ""),

    ("TC-OW-005", "One Way", "Multi-stop / extra destinations works", "Functional", "Medium",
     "Form open",
     "From: Mumbai\nTo: Pune\n+1 stop: Lonavala",
     "1. Select One Way\n2. Click Add Stop\n3. Add Lonavala\n4. Search",
     "Extra destination added to itinerary. Distance recalculated. Fare reflects multi-city",
     ""),

    ("TC-OW-006", "One Way", "Distance calculated correctly", "Functional", "High",
     "Search complete",
     "Mumbai → Pune (~150 km)",
     "1. Search Mumbai → Pune\n2. View select-car page distance",
     "Distance shows ~150 km (or as per Google API). Matches displayKms format on confirmation",
     ""),

    ("TC-OW-007", "One Way", "Surge pricing displays when applicable", "Functional", "Medium",
     "Peak time / weekend booking",
     "Saturday 6 PM",
     "1. Search peak-hour booking\n2. Look at fare breakdown",
     "Surge component shown separately if API returns surge. Total reflects surge",
     ""),

    ("TC-OW-008", "One Way", "Inclusions/exclusions visible on car card", "UI", "Medium",
     "On select-car page",
     "—",
     "1. Reach select-car\n2. Click Inclusions tab on a car",
     "Inclusions list dynamic from API. Sky-blue circular icons. INCLUDED labels removed",
     "Recent change verified"),

    ("TC-OW-009", "One Way", "Long pickup address (boundary)", "Boundary", "Low",
     "Booking details page",
     "Pickup: 200-char address",
     "1. Enter very long pickup address\n2. Try to proceed",
     "Field accepts up to API max. No truncation in UI display. Form submits cleanly",
     ""),

    ("TC-OW-010", "One Way", "Empty pickup address validation", "Negative", "High",
     "Booking details page",
     "Pickup: blank",
     "1. Leave pickup address empty\n2. Click Next",
     "Validation error: 'Pickup address required'. Cannot proceed",
     ""),

    ("TC-OW-011", "One Way", "Customer mobile must be 10 digits", "Validation", "High",
     "Booking details page",
     "Mobile: 1234",
     "1. Enter 4-digit mobile\n2. Try to proceed",
     "Error: 'Enter valid 10-digit mobile number'",
     ""),

    ("TC-OW-012", "One Way", "Email format validation on booking", "Validation", "High",
     "Booking details page",
     "Email: invalid-email",
     "1. Enter invalid email\n2. Try to proceed",
     "Error: 'Invalid email format'",
     ""),
]


# ─── 4. ROUND TRIP ───────────────────────────────────────────────
ROUNDTRIP_TESTS = [
    ("TC-RT-001", "Round Trip", "Complete round trip booking", "Functional", "Critical",
     "Logged in, wallet has balance",
     "Mumbai → Pune\nReturn next day",
     "1. Select Round Trip tab\n2. Fill from/to\n3. Pickup tomorrow\n4. Return day after\n5. Search\n6. Complete booking",
     "Booking created with both pickup + return dates. Duration = 2 days. Confirmation shows return time too",
     ""),

    ("TC-RT-002", "Round Trip", "Multi-city round trip with extra destinations", "Functional", "High",
     "Round trip selected",
     "Mumbai → Pune → Nashik → Mumbai",
     "1. Round Trip tab\n2. Add 'Nashik' as extra destination\n3. Search\n4. Select car\n5. Complete booking",
     "All cities flow into itinerary, multicityId sent in API. Confirmation shows all stops in timeline",
     ""),

    ("TC-RT-003", "Round Trip", "Return date before pickup date errors", "Negative", "High",
     "Form open",
     "Pickup: tomorrow\nReturn: today",
     "1. Set pickup tomorrow\n2. Set return today\n3. Try Search",
     "Validation error: 'Return date must be after pickup date'",
     ""),

    ("TC-RT-004", "Round Trip", "Duration auto-calculated from dates", "Functional", "Medium",
     "Round trip selected",
     "Pickup: today\nReturn: 3 days later",
     "1. Set dates 3 days apart\n2. Search\n3. Check API call",
     "duration = 4 (or correct count) sent to availability API",
     ""),

    ("TC-RT-005", "Round Trip", "Add 4+ destinations (boundary)", "Boundary", "Medium",
     "Round trip selected",
     "Mumbai + 4 stops",
     "1. Add 4 extra destinations\n2. Search",
     "All 4 stops visible. Extra destinations array correctly built. Booking creates",
     ""),

    ("TC-RT-006", "Round Trip", "Remove an extra destination", "Functional", "Low",
     "3 stops added",
     "—",
     "1. Add 3 stops\n2. Click X on middle stop\n3. Search",
     "Middle stop removed. Order maintained. Fare recalculated",
     ""),

    ("TC-RT-007", "Round Trip", "Drop address NOT required for Round Trip", "UI", "High",
     "Round trip booking flow",
     "—",
     "1. Complete round trip search → select car\n2. Reach booking page\n3. Look for drop address",
     "Drop address field hidden (not shown for Round Trip). Only pickup address required",
     "Recent change"),

    ("TC-RT-008", "Round Trip", "Pricing shows correct total with km cap", "Functional", "High",
     "Round trip with 600 km",
     "Mumbai → Pune → Mumbai",
     "1. Search round trip\n2. View select-car page\n3. Check fare calculation",
     "Total = base + km × rate. Correct distance calculation. No double-counted fields",
     ""),

    ("TC-RT-009", "Round Trip", "Premium rate applied for outstation", "Integration", "High",
     "Outstation round trip",
     "—",
     "1. Search round trip\n2. Verify network /availabilities call",
     "rate_type=premium parameter present in availability API call",
     ""),

    ("TC-RT-010", "Round Trip", "Same-day return round trip", "Functional", "Medium",
     "Round trip selected",
     "Pickup AM, Return PM same day",
     "1. Set both dates same day, time difference\n2. Complete flow",
     "Booking created with duration = 1. Both times shown correctly",
     ""),
]


# ─── 5. LOCAL ─────────────────────────────────────────────────────
LOCAL_TESTS = [
    ("TC-LOC-001", "Local", "Book Local 8hr/80km package", "Functional", "Critical",
     "Logged in",
     "City: Mumbai\nPackage: 8hr/80km",
     "1. Select Local tab\n2. Choose Mumbai\n3. Pick date/time\n4. Search\n5. On select-car: choose 8hr/80km tab\n6. Pick car\n7. Complete booking",
     "Booking created with subTripType='880'. Duration=1. No drop city",
     ""),

    ("TC-LOC-002", "Local", "Book Local 4hr/40km package", "Functional", "High",
     "Logged in",
     "City: Bangalore\nPackage: 4hr/40km",
     "1. Local tab → Bangalore\n2. Search\n3. Select 4hr/40km tab\n4. Pick car",
     "Booking with subTripType='440'. Cars filtered for 4hr package",
     ""),

    ("TC-LOC-003", "Local", "Book Local 12hr/120km package", "Functional", "High",
     "Logged in",
     "City: Delhi\n12hr/120km",
     "1. Local Delhi → search\n2. Select 12hr/120km tab",
     "Booking subTripType='12120'. Correct package selected",
     ""),

    ("TC-LOC-004", "Local", "Local search returns ALL packages (R1, R2, R3)", "Integration", "High",
     "Local search",
     "City: Mumbai",
     "1. Local tab → Mumbai\n2. Search\n3. Inspect /availabilities call",
     "API call has empty subTripType, duration=1. Response returns all 3 package tiers",
     ""),

    ("TC-LOC-005", "Local", "Drop address NOT required for Local", "UI", "High",
     "Local booking flow",
     "—",
     "1. Complete local search → select car\n2. Reach booking page",
     "Drop address field hidden. Only pickup address shown",
     "Recent change"),

    ("TC-LOC-006", "Local", "Booking summary shows just 'Local' (no 8hr/80km label)", "UI", "Low",
     "Local booking page",
     "—",
     "1. Reach booking summary\n2. Look at trip type display",
     "Shows 'Local' only. The '8hr/80km' label removed",
     "Recent change"),

    ("TC-LOC-007", "Local", "Cannot select drop city in Local", "Validation", "Medium",
     "Local tab",
     "—",
     "1. Click Local\n2. Look for drop city field",
     "No drop city field visible. Only one city selector",
     ""),

    ("TC-LOC-008", "Local", "Past pickup time blocked", "Validation", "Medium",
     "Local form open",
     "Today 06:00 (if current = 10:00)",
     "1. Local tab → today's date\n2. Try to select 06:00",
     "Past times greyed out for today. Future times allowed",
     ""),

    ("TC-LOC-009", "Local", "Filter cars by package tab on select-car", "UI", "High",
     "On select-car (local)",
     "—",
     "1. Local search → select-car\n2. Click 8hr/80km tab\n3. Click 12hr/120km tab",
     "Car list filters by package on each tab click. Prices update accordingly",
     "Recent change"),

    ("TC-LOC-010", "Local", "Premium variant (Innova) shows in local", "Functional", "Medium",
     "Local search",
     "—",
     "1. Search local Mumbai\n2. Look for Innova/premium cars",
     "Innova listed. Higher price. Fuel type/AC info visible",
     ""),
]


# ─── 6. AIRPORT PICKUP ───────────────────────────────────────────
AIRPORT_PICKUP_TESTS = [
    ("TC-AP-001", "Airport Pickup", "Complete airport pickup booking", "Functional", "Critical",
     "Logged in",
     "From: Bangalore Airport\nTo: City address",
     "1. Airport tab → Pickup from Airport\n2. Select airport\n3. Select drop city/address\n4. Date/time\n5. Search\n6. Select car\n7. Pay & confirm",
     "Booking created with subTripType='pick_airport' (NOT pickup_airport). Confirmation shows airport in pickup",
     "Recently fixed param name"),

    ("TC-AP-002", "Airport Pickup", "subTripType param is 'pick_airport' not 'pickup_airport'", "Integration", "Critical",
     "Airport pickup search",
     "—",
     "1. Switch to Airport → Pickup\n2. Search\n3. Open Network → /availabilities call",
     "Query param subTripType=pick_airport (verify exact value). Server does not return error 51001",
     "REGRESSION CHECK: Shubhendu reported earlier"),

    ("TC-AP-003", "Airport Pickup", "Airport autocomplete loads terminals", "Functional", "High",
     "Airport pickup tab",
     "Type: 'Bangalore'",
     "1. Airport pickup → click airport field\n2. Type 'Banga'",
     "Dropdown shows Bangalore International Airport with terminal options (T1, T2)",
     ""),

    ("TC-AP-004", "Airport Pickup", "Terminal selection populates terminalId", "Functional", "Medium",
     "Airport selected",
     "T1 / T2",
     "1. Select airport\n2. Pick Terminal 1\n3. Verify network call",
     "terminalId param populated in availability/booking call",
     ""),

    ("TC-AP-005", "Airport Pickup", "Drop city/address required for Pickup", "Validation", "High",
     "Airport pickup form",
     "Empty drop",
     "1. Select airport\n2. Leave drop empty\n3. Click Search",
     "Validation error: 'Drop address required'",
     ""),

    ("TC-AP-006", "Airport Pickup", "Trip type field width does not crop label", "UI", "Low",
     "Airport tab",
     "—",
     "1. Click Airport tab\n2. Open trip type dropdown",
     "Trip type field width 240px. 'Drop to Airport' / 'Pickup from Airport' labels not cropped",
     "Recent fix"),

    ("TC-AP-007", "Airport Pickup", "Switch from drop to pickup updates fields", "UI", "Medium",
     "Airport tab loaded",
     "—",
     "1. Default = Drop. Switch to Pickup\n2. Observe form fields",
     "Field labels update: From=Airport, To=City. Field positions swap correctly",
     ""),

    ("TC-AP-008", "Airport Pickup", "Pickup time validation (must be future)", "Validation", "High",
     "Airport pickup form",
     "Past time today",
     "1. Select today's date\n2. Try to pick a past time",
     "Past times not selectable. Validation error if forced",
     ""),
]


# ─── 7. AIRPORT DROP ─────────────────────────────────────────────
AIRPORT_DROP_TESTS = [
    ("TC-AD-001", "Airport Drop", "Default trip type on Airport tab is Drop", "UI", "High",
     "Click Airport tab",
     "—",
     "1. Open dashboard\n2. Click Airport tab\n3. Observe default selection",
     "Trip Type dropdown shows 'Drop to Airport' selected by default. Fields visible immediately",
     "Recent fix"),

    ("TC-AD-002", "Airport Drop", "Complete airport drop booking", "Functional", "Critical",
     "Logged in",
     "From: City\nTo: Bangalore Airport",
     "1. Airport tab (default = drop)\n2. Pickup address in city\n3. Drop = Airport\n4. Date/time\n5. Search → select car → pay",
     "Booking with subTripType='drop_airport'. Confirmation shows airport in drop",
     ""),

    ("TC-AD-003", "Airport Drop", "subTripType is 'drop_airport'", "Integration", "Critical",
     "Drop search",
     "—",
     "1. Airport drop → search\n2. Open network /availabilities",
     "subTripType=drop_airport in query params",
     ""),

    ("TC-AD-004", "Airport Drop", "Pickup address autocomplete works", "Functional", "High",
     "Airport drop form",
     "Type: 'Korama'",
     "1. Click pickup address field\n2. Type 'Korama'",
     "Autocomplete shows Koramangala suggestions from Savaari API",
     ""),

    ("TC-AD-005", "Airport Drop", "Airport list loads in drop field", "Functional", "High",
     "Drop airport selector",
     "—",
     "1. Click drop airport field\n2. Select airport",
     "Airport list loads. Selection populates field correctly",
     ""),

    ("TC-AD-006", "Airport Drop", "Empty pickup address validation", "Negative", "High",
     "Drop form",
     "Pickup: blank",
     "1. Leave pickup blank\n2. Try Search",
     "Validation error: 'Pickup address required'",
     ""),

    ("TC-AD-007", "Airport Drop", "No rate_type=premium for airport", "Integration", "Medium",
     "Airport drop search",
     "—",
     "1. Search airport drop\n2. Inspect /availabilities call",
     "No rate_type param (only outstation has premium). No destinationCity",
     ""),

    ("TC-AD-008", "Airport Drop", "Confirmation page shows correct airport in route", "UI", "Medium",
     "Booking confirmed",
     "—",
     "1. Complete drop airport booking\n2. View confirmation page",
     "Timeline shows: Pickup City → Bangalore Airport (Terminal X)",
     ""),
]


# ─── 8. CAR SELECTION ────────────────────────────────────────────
CARSELECT_TESTS = [
    ("TC-CAR-001", "Car Selection", "Car list loads with all info", "Functional", "Critical",
     "After search",
     "Mumbai → Pune",
     "1. Complete search\n2. Land on /select-car",
     "Cars show: image (left), name + 'or equivalent' (inline), pills, price, KMs, SELECT CAR button",
     "Recent redesign verified"),

    ("TC-CAR-002", "Car Selection", "Each car has 4 tabs (incl/excl/facilities/T&C)", "UI", "High",
     "On select-car",
     "—",
     "1. View any car card\n2. Click each tab",
     "Tabs visible inside card: Inclusions, Exclusions, Facilities, T&C. All clickable",
     ""),

    ("TC-CAR-003", "Car Selection", "Inclusions are dynamic from API", "Integration", "High",
     "Car card open",
     "—",
     "1. Click Inclusions tab\n2. Inspect content vs API response",
     "Inclusions list comes from API response. Decoded unicode chars (₹ shows correctly, not \\u20B9)",
     "Recent change"),

    ("TC-CAR-004", "Car Selection", "Exclusions show without cross icon", "UI", "Low",
     "Car card",
     "—",
     "1. Click Exclusions tab\n2. Observe icons",
     "Text-only exclusions. No cross icon. ON ACTUALS labels removed",
     "Recent change"),

    ("TC-CAR-005", "Car Selection", "Pill icons all sky-blue", "UI", "Low",
     "Car card",
     "—",
     "1. Look at AC/seater/luggage pills",
     "All pill icons in sky-500 color. No mixed colors",
     ""),

    ("TC-CAR-006", "Car Selection", "Modify search button opens modal", "Functional", "High",
     "On select-car page",
     "—",
     "1. Click 'Modify Search' button at top",
     "Modal opens with current trip details. From/To autocomplete fields, date/time pickers",
     ""),

    ("TC-CAR-007", "Car Selection", "Modify modal autocomplete works", "Functional", "Medium",
     "Modal open",
     "Type: 'Pune'",
     "1. Click From field\n2. Type 'Pune'",
     "Autocomplete dropdown loads PrimeNG cities via CityService",
     ""),

    ("TC-CAR-008", "Car Selection", "Modify modal time picker has 15-min steps", "UI", "Low",
     "Modal open",
     "—",
     "1. Open time picker in modal\n2. Inspect minutes",
     "15-min intervals (00, 15, 30, 45)",
     ""),

    ("TC-CAR-009", "Car Selection", "Explore Cabs button refetches", "Integration", "High",
     "Modal modified",
     "Change cities",
     "1. Modify cities/date in modal\n2. Click 'Explore Cabs'",
     "New /availabilities call fired. Car list updates live with new params. Modal closes",
     "Recent change"),

    ("TC-CAR-010", "Car Selection", "Empty result handling", "Negative", "Medium",
     "Search with no cars",
     "Remote village → city",
     "1. Search a route with no cars\n2. View select-car",
     "Empty state shown: 'No cars available' message. No crash. Modify Search still works",
     ""),

    ("TC-CAR-011", "Car Selection", "SELECT CAR button triggers booking flow", "Functional", "Critical",
     "Cars loaded",
     "—",
     "1. Click SELECT CAR on a car\n2. Wait for redirect",
     "Redirected to /booking with selected car details preserved",
     ""),

    ("TC-CAR-012", "Car Selection", "Car prices match commission/markup", "Integration", "High",
     "Logged in with markup set",
     "Markup: 10%",
     "1. Set markup 10% in settings\n2. Search → view car prices",
     "Displayed prices = API price + markup. Verify in network response",
     ""),

    ("TC-CAR-013", "Car Selection", "20 studio photos show correctly", "UI", "Low",
     "Cars loaded",
     "—",
     "1. View different car types (Sedan, SUV, Premium SUV, Tempo)",
     "Each car has appropriate studio photo. No broken images",
     ""),

    ("TC-CAR-014", "Car Selection", "KMs displayed correctly per package", "UI", "Medium",
     "Local 8hr car",
     "—",
     "1. Select local 8hr/80km\n2. View car KMs",
     "Each car shows '80 km' included for 8hr package",
     ""),

    ("TC-CAR-015", "Car Selection", "Price + SELECT CAR button center-aligned", "UI", "Low",
     "Car card",
     "—",
     "1. Visual inspection of card",
     "Price label and SELECT CAR button center-aligned vertically",
     "Recent change"),
]


# ─── 9. BOOKING DETAILS PAGE ─────────────────────────────────────
BOOKING_TESTS = [
    ("TC-BK-001", "Booking", "Page loads with selected car summary", "Functional", "Critical",
     "Car selected",
     "—",
     "1. Select car → land on /booking\n2. Observe right panel",
     "Right panel shows: car image, name, price, KMs, route, payment options",
     ""),

    ("TC-BK-002", "Booking", "Passenger name field accepts text", "Functional", "Critical",
     "On /booking",
     "Name: Rajesh Kumar",
     "1. Type passenger name\n2. Tab out",
     "Name field accepts. Validation passes. No crash",
     ""),

    ("TC-BK-003", "Booking", "Mobile field validates 10 digits", "Validation", "High",
     "Booking page",
     "Mobile: 12345 / 1234567890",
     "1. Type 5 digits → check error\n2. Type 10 digits → check OK",
     "5 digits = error. 10 digits = valid",
     ""),

    ("TC-BK-004", "Booking", "Email field validates format", "Validation", "High",
     "Booking page",
     "Email: test@example.com",
     "1. Type invalid → error\n2. Type valid → OK",
     "Email validation works as expected",
     ""),

    ("TC-BK-005", "Booking", "Country code selector loads 225 countries", "Functional", "Medium",
     "Booking page",
     "—",
     "1. Click country code dropdown\n2. Search for 'IND' or 'USA'",
     "Country code list loads. Search filters work. India shows '+91'",
     ""),

    ("TC-BK-006", "Booking", "GST checkbox auto-ticked if profile has GST", "Functional", "High",
     "Profile has GST",
     "—",
     "1. Ensure GST in profile\n2. Open booking page\n3. Look at GST section",
     "Checkbox auto-ticked. Green 'GST Applied' card shown with GSTIN",
     ""),

    ("TC-BK-007", "Booking", "GST not in profile redirects to settings", "Functional", "Medium",
     "Profile without GST",
     "—",
     "1. Remove GST from profile\n2. Open booking\n3. Tick GST checkbox",
     "Modal/redirect to /account-settings with GST section highlighted",
     ""),

    ("TC-BK-008", "Booking", "Trip Details card shows correct values", "UI", "High",
     "Booking page",
     "—",
     "1. View 'Trip Details' card on right",
     "Title is 'Trip Details' (not 'Trip Fare'). Shows route, date, KMs, fare",
     "Recent change"),

    ("TC-BK-009", "Booking", "Payment Option 1 default selected", "UI", "High",
     "Booking page",
     "—",
     "1. Open booking page\n2. Look at payment options",
     "Option 1 ('Pay any amount now') selected by default. Slider visible",
     ""),

    ("TC-BK-010", "Booking", "Slider Option 1 ranges 25-100%", "Functional", "High",
     "Option 1 selected",
     "Total: 2000",
     "1. Drag slider from 25% to 100%\n2. Observe pay-now amount",
     "25% = 500, 50% = 1000, 100% = 2000. Driver collects = total - pay-now",
     ""),

    ("TC-BK-011", "Booking", "Option 2 shows '48 hours before' text", "UI", "Medium",
     "Click Option 2",
     "—",
     "1. Click Payment Option 2\n2. Read description",
     "Description includes '48 hours before trip' for auto-deduct mention",
     "Recent change"),

    ("TC-BK-012", "Booking", "Option 3 shows '75% + 20%' (not 'deposit')", "UI", "Low",
     "Click Option 3",
     "—",
     "1. Click Payment Option 3\n2. Read description",
     "Text says '75% + 20%' (not '75% + deposit')",
     "Recent change"),

    ("TC-BK-013", "Booking", "Recommended badge has gradient + glow + pulse", "UI", "Low",
     "Option 3 visible",
     "—",
     "1. Look at Option 3's 'Recommended' badge",
     "Badge has gradient bg, glow effect, subtle pulse animation",
     "Recent change"),

    ("TC-BK-014", "Booking", "Drop address visible only for One Way", "UI", "High",
     "Different trip types",
     "—",
     "1. One Way booking → check drop field\n2. Local booking → check\n3. Round Trip → check",
     "Drop address visible only on One Way. Hidden for Local + Round Trip",
     "Recent change"),

    ("TC-BK-015", "Booking", "Pay Now amount uses current fare not stale", "Bug", "Critical",
     "Fare changed",
     "Original 3016, current 2623",
     "1. Get a booking with surge then surge clears\n2. Verify pay-now calculation",
     "Pay Now = current fare × 25% (e.g., 656). NOT 3016 × 25% = 754",
     "Recent fix - regression check"),
]


# ─── 10. PAYMENT FLOW ────────────────────────────────────────────
PAYMENT_TESTS = [
    ("TC-PAY-001", "Payment", "Wallet pay Option 1 - 25% slider", "Functional", "Critical",
     "Wallet ≥ 25% of fare",
     "Option 1 slider at 25%",
     "1. Booking page → Option 1\n2. Slide to 25%\n3. Click Pay\n4. Confirm wallet popup",
     "Wallet deducted 25%. Booking confirmed. Driver collects 75%",
     ""),

    ("TC-PAY-002", "Payment", "Wallet pay Option 1 - 100% full", "Functional", "Critical",
     "Wallet ≥ full fare",
     "Option 1 slider at 100%",
     "1. Slide to 100%\n2. Pay",
     "Full fare deducted. Driver collects 0",
     ""),

    ("TC-PAY-003", "Payment", "Wallet pay Option 2 - 25% upfront", "Functional", "Critical",
     "Wallet ≥ 25%",
     "Option 2 selected",
     "1. Click Option 2\n2. Pay",
     "25% deducted now. 75% scheduled for auto-deduct 48h before trip. Booking confirmed",
     ""),

    ("TC-PAY-004", "Payment", "Wallet pay Option 3 - full + 20% buffer", "Functional", "Critical",
     "Wallet ≥ 120% of fare",
     "Option 3 selected",
     "1. Click Option 3\n2. Pay",
     "100% fare + 20% buffer deducted. Total 120%. Confirmation shows buffer info",
     ""),

    ("TC-PAY-005", "Payment", "Wallet insufficient balance error", "Negative", "High",
     "Wallet < required",
     "Wallet: 100, fare: 2000",
     "1. Try to pay full fare with low balance",
     "Error message: 'Insufficient balance'. Booking not created. Suggest top-up",
     ""),

    ("TC-PAY-006", "Payment", "Razorpay Option 1 - 25%", "Functional", "Critical",
     "Razorpay test key configured",
     "Card: 4111 1111 1111 1111",
     "1. Click 'Pay with Razorpay'\n2. Razorpay modal opens\n3. Enter test card\n4. Complete payment",
     "Razorpay verifies. Booking confirmed. confirmation.php called",
     ""),

    ("TC-PAY-007", "Payment", "Razorpay Option 2 - 25%", "Functional", "Critical",
     "Test card",
     "Same as above",
     "1. Option 2 → Razorpay\n2. Pay 25%",
     "25% paid via gateway. Auto-deduct scheduled",
     ""),

    ("TC-PAY-008", "Payment", "Razorpay Option 3 - full upfront", "Functional", "Critical",
     "Test card",
     "—",
     "1. Option 3 → Razorpay\n2. Pay full amount",
     "Full + 20% buffer charged. Confirmed",
     ""),

    ("TC-PAY-009", "Payment", "Razorpay payment cancelled by user", "Negative", "High",
     "Razorpay modal open",
     "—",
     "1. Open Razorpay modal\n2. Click X / cancel button",
     "Modal closes. No booking confirmed. UI returns to booking page. No money deducted",
     ""),

    ("TC-PAY-010", "Payment", "Razorpay payment failure (test failure card)", "Negative", "High",
     "Test failure card",
     "Card: 4242 4242 4242 4242",
     "1. Use Razorpay test failure card\n2. Submit",
     "Failure shown by Razorpay. Booking not confirmed. User can retry",
     ""),

    ("TC-PAY-011", "Payment", "confirmation.php receives totalAmount param (DEV)", "Integration", "Critical",
     "Wallet payment",
     "—",
     "1. Make wallet payment\n2. Open Network → /confirmation.php call\n3. Inspect form data",
     "totalAmount param present with value = booking total fare (e.g., 2500)",
     "VERIFY THIS — Jibin's recent doc"),

    ("TC-PAY-012", "Payment", "confirmation.php receives bufferAmount param (DEV)", "Integration", "Critical",
     "Wallet payment",
     "—",
     "1. Make wallet payment Option 1 or 2\n2. Inspect /confirmation.php call",
     "bufferAmount param present. Value = 0 for Option 1/2. = round(price × 0.20) for Option 3",
     "VERIFY — Jibin's recent doc"),

    ("TC-PAY-013", "Payment", "Razorpay flow also sends totalAmount + bufferAmount (DEV)", "Integration", "Critical",
     "Razorpay payment",
     "—",
     "1. Make razorpay payment\n2. Inspect /confirmation.php call",
     "Both totalAmount + bufferAmount present in razorpay confirmation call too",
     "VERIFY — Jibin: 'Same params for razorpay'"),

    ("TC-PAY-014", "Payment", "DB row inserted in sv_advance_payment (DEV/QA)", "Integration", "Critical",
     "Payment complete",
     "—",
     "1. Complete a payment\n2. Backend: SELECT * FROM sv_advance_payment WHERE booking_id=X",
     "Row exists with payment_gateway = 17 (Wallet) or 16 (Razorpay), status = 2 (Success)",
     "Backend access required"),

    ("TC-PAY-015", "Payment", "DB row inserted in sv_booking_wallet_payment (DEV/QA)", "Integration", "Critical",
     "Payment complete",
     "—",
     "1. Complete payment\n2. SELECT * FROM sv_booking_wallet_payment WHERE booking_id=X",
     "Row exists. balance_paid_status = 1 if fully settled, 0 if cron pending",
     "Backend access required"),

    ("TC-PAY-016", "Payment", "settlement-payment called when fully settled", "Integration", "High",
     "Option 1 100% or Option 3 paid",
     "—",
     "1. Pay full amount\n2. Inspect Network for /booking/settlement-payment",
     "settlement-payment API called with bookingId, paymentAmount, method, transactionId",
     ""),

    ("TC-PAY-017", "Payment", "settlement-payment NOT called when deferred amount > 0", "Integration", "High",
     "Option 2 partial payment",
     "—",
     "1. Pay 25% via Option 2\n2. Inspect Network",
     "settlement-payment NOT called (cron will handle). Only confirmation.php fires",
     ""),

    ("TC-PAY-018", "Payment", "25% calculation correct on current fare", "Bug", "Critical",
     "Fare = 2623",
     "Pay Now = 656",
     "1. Booking with fare 2623\n2. Option 2 (25% upfront)\n3. Verify amount",
     "Pay Now = 656 (NOT 754 — verify no stale regularPrice floor)",
     "Recent fix"),

    ("TC-PAY-019", "Payment", "Booking ID shows on confirmation page", "UI", "Critical",
     "Payment complete",
     "—",
     "1. Pay successfully\n2. View confirmation page",
     "Booking ID prominent in hero. 7-8 digit number. Copyable",
     ""),

    ("TC-PAY-020", "Payment", "Confirmation emails sent (×2 calls)", "Integration", "Medium",
     "Payment complete",
     "—",
     "1. Pay successfully\n2. Check Network for /email_sent calls",
     "/email_sent fires twice (one for customer, one for partner)",
     ""),
]


# ─── 11. BOOKING CONFIRMATION PAGE ───────────────────────────────
CONFIRMATION_TESTS = [
    ("TC-CONF-001", "Confirmation", "Hero header with success animation", "UI", "High",
     "Booking just completed",
     "—",
     "1. Complete booking\n2. View confirmation page top",
     "Emerald gradient hero, decorative blur circles, large check icon, booking ID, 'Driver details in 1.5 hrs'",
     "Recent redesign"),

    ("TC-CONF-002", "Confirmation", "Trip Details timeline shows route", "UI", "High",
     "Confirmation page",
     "—",
     "1. Look at Trip Details card",
     "Vertical timeline with circle markers: Pickup → (via points) → Drop. Vertical line connector",
     "Recent redesign"),

    ("TC-CONF-003", "Confirmation", "KMs displays correctly (no '183 (175+8)')", "Bug", "High",
     "Round trip with breakdown",
     "183 (175+8) km",
     "1. Make booking with extra km\n2. View confirmation KMs field",
     "Shows '183 km' (just the total). Not '175 km' wrong, not '183 (175+8)' raw",
     "Recent fix - displayKms getter"),

    ("TC-CONF-004", "Confirmation", "Passenger info grid shows all fields", "UI", "High",
     "Booking complete",
     "—",
     "1. View passenger info card",
     "Grid shows: Name, Mobile (with country code), Email",
     ""),

    ("TC-CONF-005", "Confirmation", "Important T&C card with 6 bullet points", "UI", "Medium",
     "Confirmation page",
     "—",
     "1. Scroll to T&C card",
     "Amber theme card with: KM limit, CNG, Airport parking, Pickup/drop cities, Hill AC, Odometer verification",
     "Recent addition"),

    ("TC-CONF-006", "Confirmation", "Comfort & Support card with 24×7 number", "UI", "Medium",
     "Confirmation page",
     "—",
     "1. Look for support card",
     "Sky theme card with comfort tips + 24×7 support: '0 90 4545 0000' with online pulse indicator",
     "Recent addition"),

    ("TC-CONF-007", "Confirmation", "Payment Summary sticky on right", "UI", "Medium",
     "Confirmation page",
     "—",
     "1. Scroll the page\n2. Observe right panel",
     "Payment summary stays sticky as page scrolls. Shows Total, Deposit, Paid Now, Mode",
     ""),

    ("TC-CONF-008", "Confirmation", "Action buttons work (Print/WhatsApp/Bookings/New)", "Functional", "High",
     "Confirmation page",
     "—",
     "1. Click each: Print, WhatsApp, My Bookings, New Booking",
     "Print opens dialog. WhatsApp opens share. Bookings → /bookings. New → /dashboard",
     ""),

    ("TC-CONF-009", "Confirmation", "What Happens Next 4-step progression", "UI", "Low",
     "Confirmation page",
     "—",
     "1. View 'What Happens Next' card",
     "4 steps shown: Confirmed (✓), Driver Assigned, Journey, Invoice",
     "Recent addition"),

    ("TC-CONF-010", "Confirmation", "GST badge shows when GST opted in", "UI", "Medium",
     "Booking with GST",
     "GST in profile",
     "1. Book with GST\n2. View confirmation payment summary",
     "Green GST badge with GSTIN visible in payment summary",
     ""),
]


# ─── 12. WALLET ──────────────────────────────────────────────────
WALLET_TESTS = [
    ("TC-WAL-001", "Wallet", "Wallet dashboard loads", "Functional", "Critical",
     "Logged in",
     "—",
     "1. Click Wallet in nav\n2. Land on /wallet",
     "Page shows: Current balance card, top-up button, transaction history table",
     ""),

    ("TC-WAL-002", "Wallet", "Balance shows correct amount", "Functional", "Critical",
     "Wallet has balance",
     "—",
     "1. View wallet dashboard\n2. Note balance",
     "Balance in INR matches header balance and matches /wallet/balance API response",
     ""),

    ("TC-WAL-003", "Wallet", "Top-up button opens Razorpay", "Functional", "Critical",
     "Wallet page",
     "Amount: 500",
     "1. Click Top Up\n2. Enter ₹500\n3. Click Pay",
     "Razorpay modal opens with ₹500. Test key rzp_test_dsrBANLbHxlwZb used",
     ""),

    ("TC-WAL-004", "Wallet", "Top-up amount validation (min ₹100)", "Validation", "Medium",
     "Top-up form",
     "Amount: 50",
     "1. Enter ₹50 (below min)\n2. Try Pay",
     "Validation error: 'Minimum top-up ₹100' or button disabled",
     ""),

    ("TC-WAL-005", "Wallet", "Successful top-up credits balance", "Functional", "Critical",
     "Wallet balance: 1000",
     "Top-up: 500",
     "1. Top up ₹500 via test card\n2. Wait for completion",
     "Balance updates to 1500. Transaction added to history. Razorpay txn ID stored",
     ""),

    ("TC-WAL-006", "Wallet", "Failed top-up does not credit", "Negative", "High",
     "Razorpay failure card",
     "—",
     "1. Try top-up with failure test card",
     "Error shown. Balance unchanged. No transaction in history",
     ""),

    ("TC-WAL-007", "Wallet", "Transaction history loads", "Functional", "High",
     "Wallet has past transactions",
     "—",
     "1. Open wallet\n2. Scroll to history",
     "Transactions list with date, type (CREDIT/DEBIT/REFUND), amount, balance after, ref ID",
     ""),

    ("TC-WAL-008", "Wallet", "Transaction history pagination", "UI", "Medium",
     "20+ transactions",
     "—",
     "1. Scroll through history\n2. Click next/load more",
     "Pagination works. Loads next batch without errors",
     ""),

    ("TC-WAL-009", "Wallet", "Wallet token refreshes if expired", "Integration", "High",
     "Stale token",
     "—",
     "1. Wait for token expiry / force expiry\n2. Try wallet operation",
     "Token auto-refreshes. Operation succeeds without user re-login",
     ""),

    ("TC-WAL-010", "Wallet", "Booking deducts from wallet correctly", "Integration", "Critical",
     "Wallet: 5000",
     "Booking: 2000",
     "1. Make booking with wallet\n2. Check wallet balance",
     "Balance decreases by exactly the paid amount (e.g., 2000 for full pay, 500 for 25%)",
     ""),

    ("TC-WAL-011", "Wallet", "Refund credits wallet on cancel", "Functional", "High",
     "Cancelled booking",
     "—",
     "1. Cancel a paid booking\n2. Check wallet",
     "Refund amount credits back to wallet. Transaction marked REFUND",
     ""),

    ("TC-WAL-012", "Wallet", "Mock mode (Vercel) credits directly without Razorpay", "Integration", "Medium",
     "Vercel env",
     "—",
     "1. On Vercel demo, top up wallet",
     "No Razorpay modal. Direct credit. Mock data flag working",
     ""),
]


# ─── 13. ACCOUNT SETTINGS ────────────────────────────────────────
ACCOUNT_TESTS = [
    ("TC-ACC-001", "Account", "Profile loads with current values", "Functional", "Critical",
     "Logged in",
     "—",
     "1. Open /account-settings\n2. Observe form",
     "All fields populated: name, email, mobile, company, GST (if any), PAN, logo",
     ""),

    ("TC-ACC-002", "Account", "Update first name and save", "Functional", "High",
     "Settings page",
     "New name: Test",
     "1. Edit first name\n2. Click Save",
     "Update API called. Success message. New name reflects in header",
     ""),

    ("TC-ACC-003", "Account", "Mobile phone validation", "Validation", "Medium",
     "Settings page",
     "Mobile: 12345",
     "1. Enter 5-digit mobile\n2. Save",
     "Validation error: 'Enter 10-digit mobile'",
     ""),

    ("TC-ACC-004", "Account", "Add GST number (15 chars)", "Functional", "High",
     "GST empty",
     "GSTIN: 27AABCU9603R1ZM",
     "1. Click Add GST\n2. Enter valid 15-char GSTIN\n3. Save",
     "GST saved. Auto-decode shows: state, PAN, entity type, name initial",
     ""),

    ("TC-ACC-005", "Account", "Invalid GST format error", "Negative", "High",
     "Add GST",
     "GSTIN: INVALID123",
     "1. Enter invalid GSTIN\n2. Save",
     "Validation error: 'Invalid GSTIN format'",
     ""),

    ("TC-ACC-006", "Account", "GST locks after first save", "Functional", "Medium",
     "GST saved",
     "—",
     "1. Save GST\n2. Try to edit again",
     "GST field disabled. Helper text: 'Contact support to change'",
     ""),

    ("TC-ACC-007", "Account", "GST auto-decode (38 states + 10 entity types)", "Functional", "Medium",
     "Add GST",
     "Various GSTINs",
     "1. Try GSTINs from different states\n2. Verify decoded info",
     "Each state correctly identified. Entity type detected",
     ""),

    ("TC-ACC-008", "Account", "Add PAN (10 chars)", "Functional", "Medium",
     "PAN empty",
     "PAN: ABCDE1234F",
     "1. Enter PAN\n2. Save",
     "PAN saved. Validation: 5 letters + 4 digits + 1 letter",
     ""),

    ("TC-ACC-009", "Account", "Logo upload (PNG/JPG, <1MB)", "Functional", "Low",
     "Settings page",
     "Logo: 500KB PNG",
     "1. Click Upload Logo\n2. Select file\n3. Save",
     "File uploads. Preview shows. Saved to server",
     ""),

    ("TC-ACC-010", "Account", "Logo upload size limit (>1MB rejected)", "Negative", "Low",
     "Settings page",
     "Logo: 2MB",
     "1. Try uploading 2MB image",
     "Error: 'File too large. Max 1MB'",
     ""),
]


# ─── 14. MARKUP SETTINGS ─────────────────────────────────────────
MARKUP_TESTS = [
    ("TC-MRK-001", "Markup", "Markup settings page loads", "Functional", "Critical",
     "Logged in",
     "—",
     "1. Open /markup-settings\n2. View page",
     "Page shows current markup % and editable field",
     ""),

    ("TC-MRK-002", "Markup", "Current markup displayed correctly", "Functional", "High",
     "Markup set: 10%",
     "—",
     "1. View markup field",
     "Shows '10%' or '10' in input field. Matches API value",
     ""),

    ("TC-MRK-003", "Markup", "Update markup and save", "Functional", "High",
     "Editable field",
     "New markup: 15",
     "1. Change to 15\n2. Save",
     "Update API called. Success message. New value persists on reload",
     ""),

    ("TC-MRK-004", "Markup", "Markup applies to new bookings", "Integration", "High",
     "Markup: 20%",
     "—",
     "1. Set markup 20%\n2. Make a new search\n3. Compare prices vs base API",
     "Car prices = base × 1.20 (or commission added)",
     ""),

    ("TC-MRK-005", "Markup", "Negative markup blocked", "Validation", "Medium",
     "Form open",
     "Markup: -5",
     "1. Enter negative value\n2. Save",
     "Validation error: 'Markup cannot be negative'",
     ""),

    ("TC-MRK-006", "Markup", "Markup over 100% blocked or warned", "Validation", "Low",
     "Form open",
     "Markup: 150",
     "1. Enter 150\n2. Save",
     "Either capped at max or shows warning. Saved if allowed",
     ""),

    ("TC-MRK-007", "Markup", "Reset markup to 0", "Functional", "Medium",
     "Markup set",
     "Markup: 10",
     "1. Set to 0\n2. Save",
     "Value 0 saved. New bookings show base prices (no markup)",
     ""),

    ("TC-MRK-008", "Markup", "Permission check (only authorized users)", "Security", "Medium",
     "Logged in",
     "—",
     "1. Open /markup-settings",
     "Page accessible to all logged-in agents (or specific roles only — verify)",
     ""),
]


# ─── 15. REPORTS ─────────────────────────────────────────────────
REPORTS_TESTS = [
    ("TC-REP-001", "Reports", "Reports page loads", "Functional", "Critical",
     "Logged in with bookings",
     "—",
     "1. Click Reports in nav\n2. Land on /reports",
     "Page shows date range filter and bookings table",
     ""),

    ("TC-REP-002", "Reports", "Date range filter works", "Functional", "High",
     "Reports page",
     "Last 7 days",
     "1. Set date range to last 7 days\n2. Click Apply",
     "Table refreshes with bookings in that range only",
     ""),

    ("TC-REP-003", "Reports", "Booking details displayed correctly", "Functional", "High",
     "Has bookings",
     "—",
     "1. View report rows",
     "Each row shows: booking ID, date, route, customer, fare, status, payment mode",
     ""),

    ("TC-REP-004", "Reports", "Empty state when no bookings", "Functional", "Medium",
     "New agent / future-only date range",
     "Date range: future",
     "1. Set date range with no bookings\n2. View",
     "Empty state shown: 'No bookings in this period'. API returns 204 — handled correctly",
     ""),

    ("TC-REP-005", "Reports", "Sort by date works", "UI", "Medium",
     "Multiple bookings",
     "—",
     "1. Click date column header\n2. Click again",
     "Toggles ascending/descending sort",
     ""),

    ("TC-REP-006", "Reports", "Sort by amount works", "UI", "Low",
     "Multiple bookings",
     "—",
     "1. Click amount column",
     "Sorts by total amount",
     ""),

    ("TC-REP-007", "Reports", "Export to CSV (if implemented)", "Functional", "Low",
     "Has bookings",
     "—",
     "1. Click Export\n2. Save file",
     "CSV downloads with all visible columns. Encoding correct",
     "Verify if available"),

    ("TC-REP-008", "Reports", "Pagination works for large datasets", "UI", "Medium",
     "50+ bookings",
     "—",
     "1. View page 1\n2. Click Next",
     "Pagination loads next set. Page indicator updates",
     ""),
]


# ─── 16. MY BOOKINGS ─────────────────────────────────────────────
MYBOOKINGS_TESTS = [
    ("TC-MB-001", "My Bookings", "Page loads with 3 tabs", "Functional", "Critical",
     "Has bookings",
     "—",
     "1. Click My Bookings\n2. Land on /bookings",
     "3 tabs visible: Upcoming, Completed, Cancelled. Active tab highlighted",
     ""),

    ("TC-MB-002", "My Bookings", "Upcoming bookings list loads", "Functional", "High",
     "Has future bookings",
     "—",
     "1. Click Upcoming tab",
     "Bookings with future pickup dates shown. Sorted by date",
     ""),

    ("TC-MB-003", "My Bookings", "Completed bookings list loads", "Functional", "High",
     "Has past bookings",
     "—",
     "1. Click Completed tab",
     "Past bookings shown. Status = Completed",
     ""),

    ("TC-MB-004", "My Bookings", "Cancelled bookings list loads", "Functional", "High",
     "Has cancelled bookings",
     "—",
     "1. Click Cancelled tab",
     "Cancelled bookings shown with refund status",
     ""),

    ("TC-MB-005", "My Bookings", "Click booking opens details", "Functional", "High",
     "Has booking",
     "—",
     "1. Click any booking row",
     "Opens detail view or navigates to receipt page",
     ""),

    ("TC-MB-006", "My Bookings", "Cancel booking flow", "Functional", "High",
     "Has upcoming booking",
     "—",
     "1. Click Cancel on a booking\n2. Confirm in dialog\n3. Check status",
     "Booking moves to Cancelled tab. Refund processed if applicable",
     ""),

    ("TC-MB-007", "My Bookings", "Cancellation refund shows", "Functional", "Medium",
     "Cancelled booking",
     "—",
     "1. View cancelled booking",
     "Refund amount shown (if applicable). Method (wallet/bank) visible",
     ""),

    ("TC-MB-008", "My Bookings", "Empty state per tab", "UI", "Medium",
     "No bookings in a tab",
     "—",
     "1. Open a tab with no bookings",
     "Empty illustration + message: 'No upcoming bookings' (etc.)",
     ""),

    ("TC-MB-009", "My Bookings", "Search by booking ID", "Functional", "Low",
     "Many bookings",
     "Search: 12345",
     "1. Type booking ID in search\n2. Press Enter",
     "Matching booking filtered. Other rows hidden",
     "Verify if search exists"),

    ("TC-MB-010", "My Bookings", "Print receipt from booking", "Functional", "Medium",
     "Booking selected",
     "—",
     "1. Click receipt/print on a booking",
     "Print dialog opens with formatted receipt",
     ""),
]


# ─── 17. STATIC PAGES ────────────────────────────────────────────
STATIC_TESTS = [
    ("TC-STAT-001", "Static Pages", "About Us page loads", "Functional", "Low",
     "—",
     "URL: /about-us",
     "1. Navigate to /about-us",
     "Page loads with About content. No 404. Header/footer visible",
     ""),

    ("TC-STAT-002", "Static Pages", "Privacy Policy page loads", "Functional", "Low",
     "—",
     "URL: /privacy-policy",
     "1. Navigate to /privacy-policy",
     "Page loads with privacy content",
     ""),

    ("TC-STAT-003", "Static Pages", "Terms & Conditions page loads", "Functional", "Low",
     "—",
     "URL: /terms-conditions",
     "1. Navigate to /terms-conditions",
     "Page loads with T&C content",
     ""),

    ("TC-STAT-004", "Static Pages", "Contact Us page loads", "Functional", "Low",
     "—",
     "URL: /contact-us",
     "1. Navigate to /contact-us",
     "Page loads with contact info / form",
     ""),

    ("TC-STAT-005", "Static Pages", "Footer links work", "UI", "Low",
     "Any page",
     "—",
     "1. Scroll to footer\n2. Click each link",
     "All footer links navigate to correct pages",
     ""),
]


# ─── 18. CROSS-BROWSER & UI ──────────────────────────────────────
CROSSBROWSER_TESTS = [
    ("TC-UI-001", "UI", "Works in Chrome", "Compat", "Critical",
     "Chrome latest",
     "—",
     "1. Open app in Chrome\n2. Test login + booking",
     "All features work. No console errors",
     ""),

    ("TC-UI-002", "UI", "Works in Firefox", "Compat", "High",
     "Firefox latest",
     "—",
     "1. Open in Firefox\n2. Test login + booking",
     "All features work",
     ""),

    ("TC-UI-003", "UI", "Works in Edge", "Compat", "High",
     "Edge latest",
     "—",
     "1. Open in Edge\n2. Test login + booking",
     "All features work",
     ""),

    ("TC-UI-004", "UI", "Mobile Chrome (responsive)", "Compat", "High",
     "Phone or DevTools mobile",
     "—",
     "1. Open on mobile or set viewport 375px\n2. Test all flows",
     "Layout responsive. Buttons tappable. No horizontal scroll",
     ""),

    ("TC-UI-005", "UI", "Dark mode toggle works", "UI", "Medium",
     "Any page",
     "—",
     "1. Toggle dark mode (system or app toggle)",
     "All pages render correctly in dark mode. Colors invert. No unreadable text",
     ""),

    ("TC-UI-006", "UI", "Desktop 1920px layout", "UI", "Medium",
     "1920×1080 screen",
     "—",
     "1. View on 1920px monitor",
     "Layout uses full width. No empty side margins. Content centered properly",
     ""),

    ("TC-UI-007", "UI", "Tablet 768px layout", "UI", "Medium",
     "iPad / DevTools 768px",
     "—",
     "1. Set viewport 768px\n2. Test forms and lists",
     "Layout adapts. 2-col layouts may collapse to 1-col. Forms usable",
     ""),

    ("TC-UI-008", "UI", "Mobile 375px layout", "UI", "High",
     "iPhone SE / DevTools 375px",
     "—",
     "1. Set viewport 375px\n2. Test full booking flow",
     "All elements fit. Modals scrollable. Buttons touch-friendly. No overflow",
     ""),
]


# ════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, modules):
    ws = wb.create_sheet(title="Summary", index=0)

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "B2B CAB Portal — Dev Test Cases"
    cell.font = Font(bold=True, size=18, color="0EA5E9", name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    cell = ws["A2"]
    cell.value = "Module-wise dev test cases for manual verification by QA / Dev team"
    cell.font = Font(size=11, italic=True, color="64748B", name='Calibri')
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells("A3:F3")
    cell = ws["A3"]
    cell.value = "Generated: April 2026  |  Test Environment: Alpha (b2bcab.alphasavaari.com) / Vercel demo"
    cell.font = Font(size=10, color="64748B", name='Calibri')
    cell.alignment = Alignment(horizontal='center')

    # Module summary table header
    headers = ["#", "Module", "Test Count", "Critical", "High", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(bold=True, color=HEADER_TEXT, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Column widths
    widths = [6, 28, 14, 12, 10, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Module rows
    total_cases = 0
    total_crit = 0
    total_high = 0
    for i, (sheet_name, tests) in enumerate(modules, start=1):
        crit = sum(1 for t in tests if t[4] == "Critical")
        high = sum(1 for t in tests if t[4] == "High")
        row = 5 + i
        values = [i, sheet_name, len(tests), crit, high, "Pending"]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(size=10, name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(size=10, bold=True, name='Calibri')
        total_cases += len(tests)
        total_crit += crit
        total_high += high

    # Total row
    total_row = 5 + len(modules) + 1
    ws.cell(row=total_row, column=1, value="").fill = PatternFill("solid", fgColor=DARK_BG)
    cell = ws.cell(row=total_row, column=2, value="TOTAL")
    cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    cell.fill = PatternFill("solid", fgColor=DARK_BG)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    for col, val in zip([3, 4, 5, 6], [total_cases, total_crit, total_high, ""]):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=DARK_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[total_row].height = 28

    # Sign-off section
    signoff_start = total_row + 3
    ws.cell(row=signoff_start, column=1, value="Sign-off").font = Font(bold=True, size=14, color="0F172A", name='Calibri')

    signoff_rows = [
        ("Tested By", ""),
        ("Date", ""),
        ("Total Pass", ""),
        ("Total Fail", ""),
        ("Total Blocked", ""),
        ("Notes / Issues", ""),
        ("Approved By", ""),
    ]
    for i, (label, value) in enumerate(signoff_rows):
        r = signoff_start + 1 + i
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, size=10, name='Calibri')
        c1.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.fill = PatternFill("solid", fgColor=ALT_ROW)
        c2.border = border
        ws.row_dimensions[r].height = 22

    # Legend
    legend_start = signoff_start + len(signoff_rows) + 3
    ws.cell(row=legend_start, column=1, value="Legend").font = Font(bold=True, size=14, color="0F172A", name='Calibri')

    legends = [
        ("Critical", SEV_CRITICAL, "Blocker — must pass for release"),
        ("High", SEV_HIGH, "Major functionality"),
        ("Medium", SEV_MEDIUM, "Important but not blocking"),
        ("Low", SEV_LOW, "Nice to have / cosmetic"),
    ]
    for i, (label, color, desc) in enumerate(legends):
        r = legend_start + 1 + i
        c1 = ws.cell(row=r, column=1, value=label)
        c1.fill = PatternFill("solid", fgColor=color)
        c1.font = Font(bold=True, size=10, name='Calibri')
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.border = border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.font = Font(size=10, name='Calibri')
        c2.alignment = Alignment(horizontal='left', vertical='center')


def build_test_data_sheet(wb):
    ws = wb.create_sheet(title="Test Data", index=1)

    # Title
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = "Test Data — Credentials, Cities, Cards, Sample Inputs"
    cell.font = Font(bold=True, size=14, color="0EA5E9", name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    sections = [
        ("Beta Login Credentials", [
            ("Email", "bincy.joseph@savaari.com"),
            ("Password", "aMuWysE@YgAVa5aPagYR"),
            ("Agent ID", "983680"),
            ("Notes", "Use ONLY for beta/alpha. NOT production."),
        ]),
        ("Production Login (DO NOT USE FOR TESTING)", [
            ("Email", "bincy.joseph@savaari.com"),
            ("Password", "E5YJeNAgyPy#Yvajurug"),
            ("Notes", "Real bookings will trigger live SMS. AVOID."),
        ]),
        ("Razorpay Test Cards (for Razorpay flow tests)", [
            ("Success Card", "4111 1111 1111 1111"),
            ("CVV", "123"),
            ("Expiry", "Any future date"),
            ("OTP", "1234"),
            ("Failure Card", "4242 4242 4242 4242"),
            ("UPI Success", "success@razorpay"),
            ("UPI Failure", "failure@razorpay"),
        ]),
        ("Sample Cities (For Search Tests)", [
            ("Mumbai", "ID: 377"),
            ("Pune", "ID: 419"),
            ("Bangalore", "ID: 95"),
            ("Delhi", "ID: 168"),
            ("Hyderabad", "ID: 247"),
            ("Chennai", "ID: 130"),
            ("Lonavala", "Multi-stop test"),
            ("Nashik", "Round trip test"),
        ]),
        ("Sample Routes for Each Trip Type", [
            ("Outstation One Way", "Mumbai → Pune (~150 km)"),
            ("Outstation Round Trip", "Bangalore → Mysore → Bangalore (~280 km)"),
            ("Local 8hr/80km", "Mumbai (in-city)"),
            ("Local 4hr/40km", "Bangalore (in-city)"),
            ("Airport Pickup", "Bangalore Airport → MG Road"),
            ("Airport Drop", "Koramangala → Bangalore Airport"),
        ]),
        ("Sample GST Numbers (15-char GSTIN)", [
            ("Maharashtra", "27AABCU9603R1ZM"),
            ("Karnataka", "29AABCU9603R1ZJ"),
            ("Delhi", "07AABCU9603R1ZX"),
            ("Tamil Nadu", "33AABCU9603R1ZA"),
            ("Notes", "Auto-decode tests need GSTIN from different states"),
        ]),
        ("Sample PAN", [
            ("Valid", "ABCDE1234F"),
            ("Format", "5 letters + 4 digits + 1 letter"),
        ]),
        ("Sample Passenger Data", [
            ("Name", "Rajesh Kumar"),
            ("Mobile", "9876543210"),
            ("Email", "test@example.com"),
            ("Country Code", "+91 (India)"),
        ]),
        ("Booking IDs from Past Tests (for reference)", [
            ("Production (DO NOT REUSE)", "10243232, 10243266, 10243578, 10243689"),
            ("Notes", "These were accidentally created on prod. Avoid retesting"),
        ]),
        ("Database Tables (for DB verification tests)", [
            ("sv_advance_payment", "Payment records, gateway = 16 (Razorpay) / 17 (Wallet)"),
            ("sv_booking_wallet_payment", "Auto-pay queue, balance_paid_status 0/1"),
            ("sv_booking_details", "Main booking, payment_status, made_payment"),
        ]),
    ]

    row = 3
    for section_title, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.fill = PatternFill("solid", fgColor=SECTION_BG)
        cell.font = Font(bold=True, size=11, color="0F172A", name='Calibri')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.border = border
        ws.row_dimensions[row].height = 25
        row += 1

        for label, value in items:
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = Font(bold=True, size=9, name='Calibri')
            c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c1.border = border

            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = Font(size=9, name='Consolas')
            c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c2.border = border
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1  # blank line between sections

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30


def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    modules = [
        ("Auth", AUTH_TESTS),
        ("Dashboard", DASHBOARD_TESTS),
        ("One Way", ONEWAY_TESTS),
        ("Round Trip", ROUNDTRIP_TESTS),
        ("Local", LOCAL_TESTS),
        ("Airport Pickup", AIRPORT_PICKUP_TESTS),
        ("Airport Drop", AIRPORT_DROP_TESTS),
        ("Car Selection", CARSELECT_TESTS),
        ("Booking", BOOKING_TESTS),
        ("Payment", PAYMENT_TESTS),
        ("Confirmation", CONFIRMATION_TESTS),
        ("Wallet", WALLET_TESTS),
        ("Account", ACCOUNT_TESTS),
        ("Markup", MARKUP_TESTS),
        ("Reports", REPORTS_TESTS),
        ("My Bookings", MYBOOKINGS_TESTS),
        ("Static Pages", STATIC_TESTS),
        ("UI Browser", CROSSBROWSER_TESTS),
    ]

    # Build summary first (will be at index 0)
    build_summary_sheet(wb, modules)

    # Build test data sheet (index 1)
    build_test_data_sheet(wb)

    # Build module sheets
    for sheet_name, tests in modules:
        add_module_sheet(wb, sheet_name, tests)

    # Save
    output = r"C:\Users\Pranav\Downloads\B2B_Cab_Dev_Test_Cases.xlsx"
    wb.save(output)

    total = sum(len(t) for _, t in modules)
    print(f"Saved: {output}")
    print(f"Total test cases: {total}")
    print(f"Total sheets: {len(modules) + 2}")
    for sheet_name, tests in modules:
        print(f"  - {sheet_name}: {len(tests)} cases")


if __name__ == '__main__':
    main()
